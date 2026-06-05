#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для генерации шаблона Excel "Обоснование НМЦК" по 44-ФЗ
Требования: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime

def create_nmcc_template():
    # Создаем книгу и лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Обоснование НМЦК"

    # === СТИЛИ ===
    title_font = Font(name='Times New Roman', size=14, bold=True)
    header_font = Font(name='Times New Roman', size=10, bold=True)
    normal_font = Font(name='Times New Roman', size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # === ШАПКА ДОКУМЕНТА ===
    ws.merge_cells('A1:K1')
    ws['A1'] = "ОБОСНОВАНИЕ НАЧАЛЬНОЙ (МАКСИМАЛЬНОЙ) ЦЕНЫ КОНТРАКТА"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:K2')
    ws['A2'] = "(Приложение № 2 к документации о закупке)"
    ws['A2'].font = normal_font
    ws['A2'].alignment = center_align
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # === ИНФОРМАЦИЯ О ЗАКУПКЕ ===
    ws['A4'] = "Наименование закупки:"
    ws['A4'].font = normal_font
    ws.merge_cells('B4:K4')
    ws['B4'] = "{{ProcurementTitle}}"  # Плейсхолдер для замены
    ws['B4'].font = normal_font
    
    ws['A5'] = "Идентификационный код закупки (ИКЗ):"
    ws['A5'].font = normal_font
    ws.merge_cells('B5:K5')
    ws['B5'] = "{{IKZ}}"  # Плейсхолдер
    ws['B5'].font = normal_font

    # === ТАБЛИЦА КОММЕРЧЕСКИХ ПРЕДЛОЖЕНИЙ ===
    ws['A7'] = "№ п/п"
    ws['B7'] = "Наименование объекта закупки"
    ws['C7'] = "Ед. изм."
    ws['D7'] = "Предложение 1"
    ws['E7'] = "Предложение 2"
    ws['F7'] = "Предложение 3"
    ws['G7'] = "Средняя цена"
    ws['H7'] = "Количество"
    ws['I7'] = "Итого (руб.)"
    
    # Применяем стили к заголовкам
    for col in range(1, 10):
        cell = ws.cell(row=7, column=col)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill
    
    ws.row_dimensions[7].height = 40

    # === ПРИМЕР СТРОКИ (для демонстрации структуры) ===
    row_num = 8
    ws[f'A{row_num}'] = 1
    ws[f'B{row_num}'] = "{{ItemName}}"  # Плейсхолдер
    ws[f'C{row_num}'] = "{{UOM}}"
    ws[f'D{row_num}'] = "{{Price1}}"  # Цена из КП 1
    ws[f'E{row_num}'] = "{{Price2}}"  # Цена из КП 2
    ws[f'F{row_num}'] = "{{Price3}}"  # Цена из КП 3
    
    # Формула для средней цены
    ws[f'G{row_num}'] = f"=AVERAGE(D{row_num}:F{row_num})"
    
    ws[f'H{row_num}'] = "{{Quantity}}"
    
    # Формула для итога
    ws[f'I{row_num}'] = f"=G{row_num}*H{row_num}"
    
    # Применяем стили к строке данных
    for col in range(1, 10):
        cell = ws.cell(row=row_num, column=col)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    
    ws.row_dimensions[row_num].height = 25

    # === ИТОГОВАЯ СТРОКА ===
    total_row = 9
    ws.merge_cells(f'A{total_row}:H{total_row}')
    ws[f'A{total_row}'] = "ИТОГО начальная (максимальная) цена контракта:"
    ws[f'A{total_row}'].font = Font(name='Times New Roman', size=11, bold=True)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right')
    
    ws[f'I{total_row}'] = f"=SUM(I8:I{row_num})"
    ws[f'I{total_row}'].font = Font(name='Times New Roman', size=11, bold=True)
    ws[f'I{total_row}'].alignment = center_align
    
    for col in range(1, 10):
        cell = ws.cell(row=total_row, column=col)
        if col < 9:
            cell.border = thin_border
        else:
            cell.border = thin_border
            cell.fill = light_fill

    # === ИНФОРМАЦИЯ О ПОСТАВЩИКАХ ===
    info_row = 12
    ws[f'A{info_row}'] = "Информация о поставщиках, чьи коммерческие предложения использованы:"
    ws[f'A{info_row}'].font = Font(name='Times New Roman', size=10, bold=True)
    
    # Поставщик 1
    ws[f'A{info_row+1}'] = "1. Наименование:"
    ws[f'B{info_row+1}'] = "{{Provider1_Name}}"
    ws[f'A{info_row+2}'] = "   ИНН:"
    ws[f'B{info_row+2}'] = "{{Provider1_INN}}"
    ws[f'A{info_row+3}'] = "   Дата КП:"
    ws[f'B{info_row+3}'] = "{{Provider1_Date}}"
    
    # Поставщик 2
    ws[f'D{info_row+1}'] = "2. Наименование:"
    ws[f'E{info_row+1}'] = "{{Provider2_Name}}"
    ws[f'D{info_row+2}'] = "   ИНН:"
    ws[f'E{info_row+2}'] = "{{Provider2_INN}}"
    ws[f'D{info_row+3}'] = "   Дата КП:"
    ws[f'E{info_row+3}'] = "{{Provider2_Date}}"
    
    # Поставщик 3
    ws[f'G{info_row+1}'] = "3. Наименование:"
    ws[f'H{info_row+1}'] = "{{Provider3_Name}}"
    ws[f'G{info_row+2}'] = "   ИНН:"
    ws[f'H{info_row+2}'] = "{{Provider3_INN}}"
    ws[f'G{info_row+3}'] = "   Дата КП:"
    ws[f'H{info_row+3}'] = "{{Provider3_Date}}"
    
    for cell in ws[info_row:info_row+3]:
        for c in cell:
            c.font = normal_font

    # === КОЭФФИЦИЕНТ ВАРИАЦИИ ===
    cv_row = 18
    ws[f'A{cv_row}'] = "Расчет коэффициента вариации:"
    ws[f'A{cv_row}'].font = Font(name='Times New Roman', size=10, bold=True)
    
    ws[f'A{cv_row+1}'] = "Среднее квадратическое отклонение (σ):"
    ws[f'B{cv_row+1}'] = "=STDEV.S(D8:F8)"
    ws[f'A{cv_row+2}'] = "Коэффициент вариации (V):"
    ws[f'B{cv_row+2}'] = "=(B19/AVERAGE(D8:F8))*100"
    ws[f'A{cv_row+3}'] = "Однородность цен (V ≤ 33%):"
    ws[f'B{cv_row+3}'] = "=IF(B20<=33, \"ДА\", \"НЕТ\")"
    
    for i in range(cv_row, cv_row+4):
        ws[f'A{i}'].font = normal_font
        ws[f'B{i}'].font = normal_font
        ws[f'B{i}'].border = thin_border

    # === ПОДПИСИ ===
    sign_row = 25
    ws[f'A{sign_row}'] = "Ответственный за закупку:"
    ws[f'D{sign_row}'] = "_________________ / {{ResponsibleName}} /"
    
    ws[f'A{sign_row+1}'] = "Дата:"
    ws[f'D{sign_row+1}'] = "{{CurrentDate}}"
    
    for i in range(sign_row, sign_row+2):
        ws[f'A{i}'].font = normal_font
        ws[f'D{i}'].font = normal_font

    # === НАСТРОЙКИ ЛИСТА ===
    # Ширина колонок
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 5
    ws.column_dimensions['K'].width = 5

    # Печать
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.7
    ws.page_margins.bottom = 0.7

    # Сохраняем файл
    output_file = "nmcc_template.xlsx"
    wb.save(output_file)
    print(f"✅ Шаблон успешно создан: {output_file}")
    return output_file

if __name__ == "__main__":
    create_nmcc_template()
